from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.views import View
from main.models import Client, Kviz

from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
import io
import openpyxl
from openpyxl.utils import get_column_letter
from main.models import Client
from django.contrib.auth.models import User


@method_decorator(csrf_exempt, name="dispatch")
class LoginView(View):
    def post(self, request):
        user = User.objects.get(username=request.POST["username"])
        if user.check_password(request.POST["password"]):
            return HttpResponse(status=200)
        
        return HttpResponse(status=401)


@method_decorator(csrf_exempt, name="dispatch")
class AnswerView(View):
    def post(self, request: HttpRequest):
        data = request.POST
        kviz = request.POST.get("kviz")
        kviz = Kviz.objects.get(id=kviz)
        values = [i for i in data.get("value").split(";")]
        fields = [i for i in data.get("field").split(";")]

        print(values)
        print(fields)

        for field, value in zip(fields, values):
            if value == "true" or value == "on":
                value = True
            if value == "false":
                value = False
            setattr(kviz, field, value)
        kviz.save()
        return HttpResponse(status=201)


@method_decorator(csrf_exempt, name="dispatch")
class CreateKvizView(View):
    def post(self, request: HttpRequest):
        data = request.POST
        print(data)

        client = get_object_or_404(Client, id=data['client_id'])

        kviz_count = Kviz.objects.filter(client=client).count()

        kviz = Kviz.objects.create(client=client, count=kviz_count+1)

        return JsonResponse({"kviz": kviz.id})


@method_decorator(csrf_exempt, name="dispatch")
class ClientView(View):
    def post(self, request: HttpRequest):
        data = request.POST
        print(data)

        client =  Client.objects.filter(user_id=data["user_id"], messanger=data["messanger"]).first()
        if client is None:
            client = Client.objects.create(
                user_id=data["user_id"],
                messanger=data["messanger"],
                messanger_name=data.get("messanger_name"),
                nickname=data.get("nickname")
            )

        return JsonResponse({"client": client.id})

    def get(self, request: HttpRequest):
        data = request.POST
        print(data)

        client =  Client.objects.filter(user_id=data["user_id"], messanger=data["messanger"]).first()
        if client is None:
            client = Client.objects.create(
                user_id=data["user_id"],
                messanger=data["messanger"],
                messanger_name=data.get("messanger_name"),
                nickname=data.get("nickname")
            )

        return JsonResponse({"client": client.id})


class GetClientsExcel(View):
    def get(self, request):
        workbook = openpyxl.Workbook()

        sheet = workbook.active

        a = [chr(i) for i in range(ord("A"), ord("Z")+1)]

        titles_letters = a[:] + [i+j for i in a for j in a]

        titles = [
            "ТГ/ВК", "Имя в ТГ/ВК", "Имя, которое указал пользователь", "Никнейм", "User ID",
            "Телефон", "Уникальный ключ", "Дата и время", "Что интересует", 
            "Профессия", "Опыт работы", "Год рождения", "Минимальная зарплата",
            "Гражданство", "Место жительства", "О своих профессиональных навыках",
            "Ответственный", "Инструмент", "Есть напарник", "Бригада", "Авто",
            "Оценка объекта", "СЗ/ИП", "Другие преимущества", "НАКСТ", "Монтаж окожушки",
            "Монтаж фольг. цилиндров", "Монтаж K-Flex", 
            "Монтаж ППУ скорлупы", "Монтаж пеностекла", 
            "Изготовление окожушки", "Напыление ППУ", "Другие навыки",
            "Трубопроводы", "Воздуховоды", "Резервуары", "Оборудование",
            "Что другое умеет изолировать", "Ручная электродуговая", "Полуавтоматическая",
            "Аргоновая", "Газовая", "Плазменная", "Сварка в среде газа",
            "Другие виды сварки", "Простые конструкции", "Каркасы зданий",
            "Балки, колонны", "Трубопроводы", "Энергетические котлы",
            "Оборудование", "Другие конструкции", "Читаю чертежи и схемы",
            "Знаю нормы и стандарты", "Провожу гидроиспытания", "Организаторские навыки",
            "Работа с инструментами", "Другие навыки", "ОПФ",
            "Направление деятельности", "Регион/Город", "Цена контрактов", "Численность персонала",
            "Интересуют специалисты", "Специалисты в регоне",
            "Организация Форма 2", "e-mai Форма 2", "Сообщение Форма 2", "Есть основная работа",
            "Есть основная работа (Другое)", "Интересует постоянная работа",
            "Интересует временная работа", "Интересует работа без оформления", 
            "Интересует график 5/2", "Интересует сменный график", "Интересует ежедневная оплата", 
            "Интересует сдельная оплата", "Интересует любая работа",
            "Какая работа Вас интересует? (Другое)", "Звонок", "Вконтакте", 
            "Телеграм", "Ватсап", "регион объекта", "регион объекта (Другое)",
            "Штукатурные работы", "Малярные работы", "Укладка плитки", "Монтаж гипсокартона",
            "Оклейка обоями", "Установка дверей и окон", "Монтаж потолков",
            "Комплексная отделка", "Отделочник (другое)", "Метро"
        ]

        fields = [
            "messanger", "messanger_name", "name", "nickname", "user_id",
            "phone", "unique_key", "created_at_tag", "interes", 
            "profession", "work_experience", "birth_year", "min_zp",
            "citizenship", "residence", "skills", "responsible", "instrument",
            "partner", "brigada", "car", "rate_work", "have_ip",
            "another_skills", "NAKCT", "installation_window_frame",
            "installation_foil_cylinders", "installation_kflex", 
            "installation_PPU_shell", "foam_glass_installation", 
            "manufacture_shell", "spraying_PPU", "another_insulator_skills",
            "pipelines", "airducts", "reservoirs", "equipment",
            "another_can_installer_skills", "manual_electric_arc", "semiautomatic",
            "argon", "gasfired", "plasma", "welding_gas_environment",
            "another_welding", "can_weld_simple_constructions", "can_weld_building_frames",
            "can_weld_columns", "can_weld_pipelines", "can_weld_energy_boilers",
            "can_weld_equipment", "can_weld_another", "read_drawings_diagrams",
            "know_norms_standards", "conducting_hydrotests", "organizational_skills",
            "working_with_tools", "another_installer_skills", "OPF",
            "activity_direction", "activity_region", "contract_price", "ceh_count",
            "interest_professions", "region_professions",
            "organization_form2", "email_form2", "message_form2", "has_work",
            "has_work_another", "permanent_job",
            "temporary_job", "job_without_registration", "schedulefivetwo",
            "shift_schedule", "dayly_pay", "piecework_payment", "any_job", "work_another",
            "call", "vk", "tg", "whatsapp", "object_region", "another_object_region",
            "plastering_works", "painting_work", "laying_tiles",
            "installation_drywall", "wallpapering", "installation_doors_and_windows",
            "ceiling_installation", "comprehensive_finishing", "another_finisher", "metro"
        ]
    
        kvizes = Kviz.objects.select_related("client").order_by("-id").all()

        for field, letter in zip(titles, titles_letters):
            sheet[letter + '1'] = field

        for row, kviz in enumerate(kvizes):
            for ind, field in enumerate(fields):
                val = getattr(kviz, field)
                if isinstance(val, bool):
                    val = {
                        True: 'да',
                        False: ''
                    }[val]
                sheet[titles_letters[ind] + str(row+2)] = val

        for col in range(1, len(fields)+1):
            sheet.column_dimensions[get_column_letter(col)].width = 15

        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = 'attachment; filename="Опросы.xlsx"'

        return response
    

class GetKvizCountView(View):
    def get(self, request: HttpRequest):
        data = request.GET
        print(data)

        client = get_object_or_404(Client, id=data['client_id'])

        kviz_count = Kviz.objects.filter(client=client).count()

        return JsonResponse({"kviz_count": kviz_count})
