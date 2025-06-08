from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views import View
from main.forms import FORMS
from main.models import Client, Kviz
from django.views.generic import TemplateView

from seo.models import IndexPage
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt


class Index(TemplateView):
    template_name = 'main/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        kviz_id = self.request.GET.get("kviz")
        kviz = Kviz.objects.get(id=int(kviz_id))

        context["kviz"] = kviz
        context["settings"] = IndexPage.objects.first()

        context["forms"] = FORMS

        return context


class IndexVK(TemplateView):
    template_name = 'main/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #kviz_id = self.request.GET.get("kviz")
        #kviz = Kviz.objects.get(id=int(kviz_id))

        #context["kviz"] = kviz
        context["settings"] = IndexPage.objects.first()

        context["forms"] = FORMS

        return context


@method_decorator(csrf_exempt, name="dispatch")
class AnswerView(View):
    def post(self, request: HttpRequest):
        data = request.POST
        kviz = request.POST.get("kviz")
        kviz = Kviz.objects.get(id=kviz)
        values = [i for i in data.get("value").split(",")]
        fields = [i for i in data.get("field").split(",")]

        print(values)
        print(fields)

        for field, value in zip(fields, values):
            if value == "true":
                value = True
            if value == "false":
                value = False
            setattr(kviz, field, value)
        kviz.save()
        return HttpResponse(status=201)
    

@method_decorator(csrf_exempt, name="dispatch")
class CreateKvizView(View):
    def post(self, request: HttpRequest):
        data = request.POST
        if data["messanger"] == "tg":

            client = Client.objects.filter(messanger_name=data["username"], messanger=data["messanger"]).first()
            if client:
                old_kviz = Kviz.objects.filter(client=client).last()

                kviz = Kviz.objects.create(client=client, count=old_kviz.count+1)
            else:
                client = Client.objects.create(
                    messanger="tg",
                    messanger_name=data["username"]
                )
                    
                kviz = Kviz.objects.create(client=client, count=1)

        return JsonResponse({"kviz": kviz.id})


import io

class GetClientsExcel(View):
    def get():
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from main.models import Client


        # Создаем новую книгу
        workbook = openpyxl.Workbook()

        # Получаем активный лист (по умолчанию 'Sheet')
        sheet = workbook.active

        #sheet['B1'] = 'Возраст'
        #sheet['C1'] = 'Город'

        sheet['A2'] = 'Иван'
        sheet['B2'] = 30
        sheet['C2'] = 'Москва'

        sheet['A3'] = 'Мария'
        sheet['B3'] = 25
        sheet['C3'] = 'Санкт-Петербург'

        a = [chr(i) for i in range(ord("A"), ord("Z")+1)]

        titles_letters = a[:] + [i+j for i in a for j in a]

        titles = ["ТГ/ВК", "Имя в ТГ/ВК", "Имя, которое указал пользователь", "Никнейм", "User ID", 
                  "Телефон", "Уникальный ключ", "Дата и время", "Что интересует", "Профессия", 
                  "Опыт работы", "Год рождения", "Минимальная зарплата", "Гражданство", "Место жительства", 
                  "О своих профессиональных навыках", "Ответственный", "Инструмент", "Есть напарник", 
                  "Бригада", "Авто", "Оценка объекта", "СЗ/ИП", "Другие преимущества" "НАКСТ", "Монтаж окожушки",
                  "Монтаж фольг.", "Цилиндров", "Монтаж K-Flex", "Монтаж ППУ скорлупы", "Монтаж пеностекла", 
                  "Изготовление окожушки", "Напыление ППУ",	"Другие навыки", "Трубопроводы", "Воздуховоды", 
                  "Резервуары", "Оборудование", "Что другое умеет изолировать", "Ручная электродуговая", 
                  "Полуавтоматическая", "Аргоновая", "Газовая", "Плазменная", "Сварка в среде газа",
                  "Другие виды сварки", "Простые конструкции", "Каркасы зданий", "Балки, колонны",
                   "Трубопроводы", "Котлы", "Оборудование", "Другие конструкции", "Читаю чертежи и схемы", 
                   "Знаю нормы и стандарты", "Провожу гидроиспытания", "Организаторские навыки",
                    "Работа с инструментами", "Другие навыки", "ОПФ", "Направление деятельности", "Регион/Город",
                    "Цена контрактов", "Численность персонала", "Интересуют специалисты", "Организация Форма 2",
                  "e-mai Форма 2", "Сообщение Форма 2", "Есть основная работа", "пятидневка",
                    "Есть основная работа, сменный график", "Свободный график",	"Нет работы, только подработки",
                    "Нет основной работы", "У вас есть основная работа? (Другое)",	"Интересует постоянная работа",
                    "Интересует временная работа", "Интересует работа без оформления", "Интересует график 5/2",
                    "Интересует сменный график", "Интересует ежедневная оплата", "Интересует сдельная оплата",
                    "Интересует любая работа", "Какая работа Вас интересует? (Другое)", "Звонок", "Ватсап",
                    "Телеграм", "Вконтакте"
                ]
        
        for field, letter in zip(titles, titles_letters):
            sheet[letter + '1'] = field

        # Форматирование заголовков
        #header_font = Font(bold=True, color="FFFFFF") # Белый цвет текста
        #header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid") # Синий цвет фона
        #header_alignment = Alignment(horizontal="center", vertical="center")

        #for col in range(1, 4):  # A, B, C
        #    cell = sheet.cell(row=1, column=col)
        #    cell.font = header_font
        #    cell.fill = header_fill
        #    cell.alignment = header_alignment
        #    sheet.column_dimensions[get_column_letter(col)].width = 15 # Устанавливаем ширину столбца

        # Автоматическая ширина столбцов (не всегда идеально, может потребоваться настройка)
        # for column_cells in sheet.columns:
        #     length = max(len(str(cell.value)) for cell in column_cells)
        #     sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2


        # Сохраняем файл
        #workbook.save('example.xlsx')

        # 2. Сохраняем Excel-файл в байтовый поток в памяти (используем io.BytesIO)
        excel_file = io.BytesIO()
        workbook.save(excel_file)  # Сохраняем в байтовый поток
        excel_file.seek(0)  # Перематываем указатель в начало потока

        # 3. Создаем HTTP-ответ
        response = HttpResponse(
            excel_file.read(),  # Читаем байты из потока
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # MIME-тип для XLSX
        )

        # 4. Устанавливаем HTTP-заголовок Content-Disposition, чтобы указать имя файла для скачивания
        response['Content-Disposition'] = 'attachment; filename="example.xlsx"'  # Укажите желаемое имя файла

        return response